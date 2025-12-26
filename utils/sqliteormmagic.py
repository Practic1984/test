# python3.12.3
"""
Скрипт позволяет получить доступ к базе данных SQlite3. 
через функцию, что удобнее, чем 
синтаксис прямых SQL-запросов. По вопросам и 
"""
import sqlite3
import pandas as pd
import os
from sqlite3 import Error
import logging
from utils import user_sql_query
from utils import admin_sql_query
from config import config

from utils import other
logging.basicConfig(level=logging.ERROR, format="%(asctime)s - %(levelname)s - %(message)s")

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_connection(path):
    """
    Создаёт подключение к базе данных SQLite.
    :param path: str - путь к файлу базы данных.
    :return: объект подключения.
    """
    connection = None
    try:
        connection = sqlite3.connect(path)
        print("Connection to SQLite DB successful")
    except Error as e:
        logging.error(f"SQL error: {e}")
 
    return connection

def execute_query(connection, query, params=[]):
    """
    Выполняет запрос к базе данных SQLite (INSERT, UPDATE, DELETE и т. д.).
    :param connection: объект подключения к базе данных.
    :param query: str - SQL-запрос.
    :param params: list - параметры запроса (по умолчанию пустой список).
    :return: результат выполнения запроса (если есть).
    """
    res = None
    cursor = connection.cursor()
    try:

        if len(params) > 0:

            cursor.execute(query, params)
            # res = cursor.fetchone() # fetchall()
        else:
            cursor.execute(query)
            res = cursor.fetchall()
        connection.commit()
        print("Query executed successfully")

    except Error as e:
        logging.error(f"Connection error: {e}")

    return res

def execute_query_select(connection, query, params=[]):
    """
    Выполняет SELECT-запрос к базе данных SQLite и возвращает результат.
    :param connection: объект подключения к базе данных.
    :param query: str - SQL-запрос.
    :param params: list - параметры запроса (по умолчанию пустой список).
    :return: список кортежей с результатами запроса.
    """
    res = None
    cursor = connection.cursor()
    try:
        cursor.execute(query, params)
        res = cursor.fetchall()

        connection.commit()
        print("Query executed successfully")
    except Error as e:
        logging.error(f"Connection error: {e}")

    return res

class SQLiteDB():

    def __init__(self, DBNAME):
        """
        Инициализирует объект базы данных.
        :param DBNAME: str - путь к файлу базы данных.
        """
        self.DBNAME = DBNAME

    def create_table(self, create_query: str, insert_query: str = None, params: list = None):
        """
        Универсальный метод для создания таблицы и добавления данных.
        
        Args:
            create_query (str): SQL-запрос для создания таблицы.
            insert_query (str): SQL-запрос для вставки данных (опционально).
            params (list): Параметры для вставки данных (опционально).
        """
        with create_connection(self.DBNAME) as connection:
            execute_query(connection, create_query)
            if insert_query and params:
                execute_query(connection, insert_query, params)     


    def find_elements(self, table_name: str, filters: dict) -> list[dict]:
        """
        Универсальный метод для поиска записей по фильтрам.
        
        Args:
            table_name (str): Название таблицы.
            filters (dict): Словарь фильтров {имя_колонки: значение}.
        
        Returns:
            list[dict]: Список найденных строк в виде словарей.
        
        Raises:
            ValueError: Если таблица или колонка не разрешены.
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")

        # Проверяем, что все ключи фильтров являются допустимыми идентификаторами
        for column in filters.keys():
            if not column.isidentifier():
                raise ValueError(f"Недопустимое название колонки: {column}")

        # Формируем запрос
        conditions = " AND ".join([f"{col} = ?" for col in filters.keys()])
        query = f"SELECT * FROM \"{table_name}\" WHERE {conditions}"

        with create_connection(self.DBNAME) as connection:
            connection.row_factory = sqlite3.Row
            cursor = connection.cursor()
            cursor.execute(query, list(filters.values()))
            rows = cursor.fetchall()

        return [dict(row) for row in rows]

   
    def check_table(self, table: str):
        """
        Проверяет существование таблицы в базе данных.
        :param table: str - название таблицы.
        :return: bool - True, если таблица существует, иначе False.
        """
        connection = create_connection(self.DBNAME)
        cursor = connection.cursor()
        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table}';")
        res = cursor.fetchone()
        connection.close()
        if res: return True
        else:
            return False
        
    def check_user_on_table(self, table: str, from_user_id: int):
        """
        Check user on table by from_user_id
        """
        connection = create_connection(self.DBNAME)
        cursor = connection.cursor()
        cursor.execute(f"SELECT from_user_id FROM {table} WHERE from_user_id=?;", (from_user_id,))
        res = cursor.fetchone()
        connection.close()
        if res: 
            return True
        else:
            return False    
    
        
    def get_row_by_user_id(self, table_name: str, from_user_id: int):
        """
        Universal function for reading a row from any table using from_user_id.

        Args:
            table_name (str): The name of the table to query.
            from_user_id (int): The user ID to filter by.

        Returns:
            dict | None: The row as a dictionary if found, else None.
        """
        query = f"SELECT * FROM {table_name} WHERE from_user_id = ?"
        with create_connection(self.DBNAME) as connection:
            connection.row_factory = sqlite3.Row
            cursor = connection.cursor()
            cursor.execute(query, [from_user_id])
            row = cursor.fetchone()
            return dict(row) if row else None


    def get_table_report( self, message, table):
        connection = create_connection(self.DBNAME)
        query = admin_sql_query.get_table_rows.format(table=table)
        all_records = pd.read_sql_query(query, connection)
        len_of_records = len(all_records) if not all_records.empty else 0
        os.makedirs("./reports", exist_ok=True) 
        path = f'./reports/report_{table}_{message.from_user.id}.xlsx'       
        all_records.to_excel(path, index=False)
        connection.close()
        
        return path, len_of_records           

    
    def get_all(self, table: str) -> list[dict]:
        """
        Возвращает все строки из таблицы в виде списка словарей.
        :param table: str - название таблицы
        :return: list[dict]
        """
        if table not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table}")

        query = f"SELECT * FROM {table}"

        with create_connection(self.DBNAME) as connection:
            cursor = connection.cursor()
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

        return [dict(zip(columns, row)) for row in rows]

    def get_push_ca(self, table: str, leads: str) -> list[int]:
        """
        Возвращает список from_user_id из указанной таблицы с фильтром по leads.
        
        Args:
            table: Название таблицы (worker/customer_worker/customer_qualified)
            leads: Значение для фильтрации ('yes'/'no')
            
        Returns:
            Список user_id (int)
            
        Raises:
            ValueError: Если таблица не найдена в ALLOWED_TABLES
            sqlite3.Error: При ошибках работы с БД
        """
        # Проверка допустимости таблицы
        if table not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table}")

        # Проверка допустимости значения leads
        if leads not in ('yes', 'no'):
            raise ValueError("Недопустимое значение leads. Допустимо: 'yes' или 'no'")

        # Формируем запрос с экранированием имени таблицы
        query = f"SELECT from_user_id FROM \"{table}\" WHERE leads = ?"
        
        try:
            with create_connection(self.DBNAME) as conn:
                # Используем cursor вместо pandas для простой выборки
                cursor = conn.cursor()
                cursor.execute(query, (leads,))
                
                # Возвращаем список int напрямую
                return [row[0] for row in cursor.fetchall()]
                
        except sqlite3.Error as e:
            logging.error(f"Ошибка при выборке push CA из таблицы {table}: {str(e)}")
            raise

    def get_all_users(self) -> list:
        """
        Получает список всех пользователей.
        :return: list - список user_id.
        """
        query = admin_sql_query.get_all_users

        with create_connection(self.DBNAME) as connection:
            all_records_users = pd.read_sql_query(query, connection)

        # Проверяем, есть ли колонка 'from_user_id', иначе возвращаем пустой список
        if 'from_user_id' not in all_records_users.columns or all_records_users.empty:
            return []

        return all_records_users['from_user_id'].astype(int).tolist()
    
    def find_table_or_column(self, table_name: str, column_name: str) -> list[dict]:
        """
        Ищет все значения по указанным колонкам в таблице.
        :param table_name: str - название таблицы
        :param column_name: str - название колонки или несколько колонок через запятую
        :return: list[dict] - список словарей с данными
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")

        # Проверяем, что column_name содержит только разрешённые символы (буквы, цифры, запятая)
        if not all(part.isidentifier() or part == '*' for part in column_name.replace(" ", "").split(",")):
            raise ValueError("Недопустимое название колонки")

        query = f"SELECT {column_name} FROM {table_name}"

        with create_connection(self.DBNAME) as connection:
            connection.row_factory = sqlite3.Row  # Позволяет получать результаты как словари
            result = execute_query_select(connection, query=query, params=[])

        return [dict(row) for row in result] if result else []
    
    def find_elements_by_keywords(
            self, 
            table_name: str, 
            **keyword_pairs: str
        ) -> list[dict]:
        """
        Ищет записи по произвольному количеству пар ключ-колонка.

        :param table_name: str - имя таблицы
        :param keyword_pairs: произвольное количество пар в формате column_name=key_value
        :return: list[dict] - список найденных строк
        # Новый вызов с двумя ключами
        result = self.find_elements_by_keywords(
            "users", 
            name="John", 
            email="john@example.com"
        )
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")

        # Проверяем все переданные названия колонок
        for column_name in keyword_pairs.keys():
            if not column_name.isidentifier():
                raise ValueError(f"Недопустимое название колонки: {column_name}")

        # Если условия не переданы, возвращаем все записи
        if not keyword_pairs:
            query = f"SELECT * FROM {table_name}"
            params = []
        else:
            # Формируем условия WHERE
            conditions = [f"{col} = ?" for col in keyword_pairs.keys()]
            where_clause = " AND ".join(conditions)
            
            query = f"""
                SELECT * FROM {table_name} 
                WHERE {where_clause}
            """
            params = [str(value) for value in keyword_pairs.values()]

        with create_connection(self.DBNAME) as connection:
            connection.row_factory = sqlite3.Row
            result = execute_query_select(connection, query=query, params=params)

        return [dict(row) for row in result] if result else []

    def get_last_inserted_id(self, table_name: str) -> int:
        """
        Возвращает ID последней записи в указанной таблице.
        :param table_name: str - имя таблицы
        :return: int - ID последней записи или None если таблица пуста
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")

        query = f"SELECT seq FROM sqlite_sequence WHERE name = ?"
        params = [table_name]

        with create_connection(self.DBNAME) as connection:
            result = execute_query_select(connection, query=query, params=params)
            
            # Если в sqlite_sequence нет записи для этой таблицы,
            # попробуем получить максимальный ID напрямую из таблицы
            if not result:
                query = f"SELECT MAX(rowid) as last_id FROM {table_name}"
                result = execute_query_select(connection, query=query)
        
        return result[0][0] if result else None


    def upd_element_in_column(self, table_name: str, upd_column_name: str, new_value: str, key_column_name: str, key_value: str):
        """
        Обновляет данные в таблице.
        :param table_name: str - название таблицы
        :param upd_column_name: str - колонка, которую обновляем
        :param new_value: str - новое значение для установки
        :param key_column_name: str - колонка, по которой ищем
        :param key_value: str - значение для поиска
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")

        if not (upd_column_name.isidentifier() and key_column_name.isidentifier()):
            raise ValueError("Недопустимое название колонки")

        query = f"""
            UPDATE {table_name}
            SET {upd_column_name} = ?
            WHERE {key_column_name} = ?
        """

        with create_connection(self.DBNAME) as connection:
            execute_query(connection, query=query, params=[new_value, key_value])

    def update_multiple_columns_by_user_id(
        self, 
        table_name: str, 
        from_user_id: int, 
        updates: dict
    ) -> bool:
        """
        Обновляет несколько колонок в таблице по from_user_id.
        
        Args:
            table_name: Название таблицы
            from_user_id: ID пользователя для поиска записи
            updates: Словарь с обновлениями {имя_колонки: новое_значение}
            
        Returns:
            bool: True если обновление прошло успешно, False при ошибке
            
        Raises:
            ValueError: При недопустимых параметрах
        """
        # Проверка допустимости таблицы
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")
        
        # Проверка наличия обновлений
        if not updates:
            raise ValueError("Словарь обновлений не может быть пустым")
        
        # Проверка названий колонок
        for column in updates.keys():
            if not column.isidentifier():
                raise ValueError(f"Недопустимое название колонки: {column}")
        
        # Формируем SET-часть запроса
        set_clause = ', '.join([f"{col} = ?" for col in updates.keys()])
        params = list(updates.values()) + [from_user_id]
        
        query = f"""
            UPDATE {table_name}
            SET {set_clause}
            WHERE from_user_id = ?
        """
        
        try:
            with create_connection(self.DBNAME) as connection:
                cursor = connection.cursor()
                cursor.execute(query, params)
                connection.commit()
                return cursor.rowcount > 0
                
        except Error as e:
            logging.error(f"Ошибка при обновлении записи {from_user_id} в таблице {table_name}: {e}")
            return False

    def upd_element_by_filters(self, table_name: str, upd_column_name: str, new_value: str, filters: dict):
        """
        Обновляет колонку в таблице по нескольким условиям (AND).
        :param table_name: str - название таблицы
        :param upd_column_name: str - колонка, которую обновляем
        :param new_value: str - новое значение
        :param filters: dict - словарь условий (key: column, value: значение)
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")

        if not upd_column_name.isidentifier():
            raise ValueError("Недопустимое название колонки")

        for k in filters:
            if not k.isidentifier():
                raise ValueError("Недопустимое название фильтра")

        where_clause = ' AND '.join([f"{k} = ?" for k in filters])
        params = [new_value] + list(filters.values())

        query = f"""
            UPDATE {table_name}
            SET {upd_column_name} = ?
            WHERE {where_clause}
        """

        with create_connection(self.DBNAME) as connection:
            execute_query(connection, query=query, params=params)

    def get_element_by_filters(self, table_name: str, filters: dict) -> dict:
        """
        Получает одну запись из таблицы по указанным фильтрам.
        
        :param table_name: str - имя таблицы
        :param filters: dict - словарь условий (ключ - имя колонки, значение - значение для фильтрации)
        :return: dict - найденная запись или None, если запись не найдена
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")
        
        if not filters:
            raise ValueError("Фильтры не могут быть пустыми")
        
        # Проверяем названия колонок
        for column in filters:
            if not column.isidentifier():
                raise ValueError(f"Недопустимое название колонки: {column}")
        
        # Формируем WHERE-часть запроса
        where_clause = ' AND '.join([f"{k} = ?" for k in filters])
        params = list(filters.values())
        
        query = f"SELECT * FROM {table_name} WHERE {where_clause} LIMIT 1"
        
        with create_connection(self.DBNAME) as connection:
            connection.row_factory = sqlite3.Row  # Для возврата результатов в виде словаря
            cursor = connection.cursor()
            cursor.execute(query, params)
            result = cursor.fetchone()
        
        return dict(result) if result else None
    
    def delete_table(self, table):
            """
            Удаляет все записи из указанной таблицы.
            :param table: str - имя таблицы.
            """
            connection = create_connection(self.DBNAME)
            query = f"""
            DELETE FROM "{table}"
                    """ 
            execute_query(connection, query=query, params=[])
            connection.close() 

    def delete_row(self, table:str, key_name:str, column_name:str):
            """
            Удаляет строку из таблицы по указанному значению.
            :param table: str - имя таблицы.
            :param key_name: str - значение для удаления.
            :param column_name: str - имя колонки.
            """
            connection = create_connection(self.DBNAME)
            query = f"""
            DELETE FROM "{table}"
            WHERE {column_name} = ?
                    """ 
            execute_query(connection, query=query, params=[key_name])
            connection.close() 


    def ins_unique_row(self, table_name: str, values: dict) -> None:
        """
        Функция вставки данных в таблицу с обработкой уникальных значений
        (использует INSERT OR IGNORE для колонок с UNIQUE constraint)
        
        :param table_name: str - имя таблицы
        :param values: dict - словарь с данными для вставки, где:
                        ключ - название колонки,
                        значение - значение для вставки
        Пример вызова:
        db_users.ins_unique_row(
            table_name='bids',
            values={
                'service': 'cleaning',
                'max_count_workers': 3,
                'date_work': '2023-01-01',
                'comments': 'Urgent'
            }
        )
        """
        if not values:
            raise ValueError("Словарь значений не может быть пустым")
        
        # Подготовка параметров запроса
        columns = ', '.join(values.keys())
        placeholders = ', '.join(['?'] * len(values))
        params = list(values.values())
        
        query = f"""
        INSERT OR IGNORE INTO {table_name} ({columns})
        VALUES ({placeholders})
        """
        
        with create_connection(self.DBNAME) as connection:
            execute_query(connection=connection, query=query, params=params)



    def append_to_cell(self, table: str, column: str, value: str, key_column: str, key_value: any):
        """
        Дописывает данные в ячейку таблицы SQLite.
        
        :param table: str - имя таблицы
        :param column: str - имя колонки, куда дописывать
        :param value: str - строка для добавления
        :param key_column: str - имя колонки, по которой ищем строку
        :param key_value: any - значение для поиска строки
        """
        query = f"""
        UPDATE {table}
        SET {column} = COALESCE({column}, '') || ?
        WHERE {key_column} = ?
        """
        with create_connection(self.DBNAME) as connection:
            execute_query(connection, query, [value, key_value])

    def get_last_row_by_user_id(self, table_name: str, from_user_id: int) -> dict | None:
        """
        Возвращает последнюю запись из таблицы по from_user_id.

        :param table_name: str - имя таблицы
        :param from_user_id: int - ID пользователя
        :return: dict - последняя запись или None, если не найдено
        """
        if table_name not in config.ALLOWED_TABLES:
            raise ValueError(f"Недопустимое название таблицы: {table_name}")

        query = f"""
        SELECT * FROM {table_name}
        WHERE from_user_id = ?
        ORDER BY id DESC LIMIT 1
        """

        with create_connection(self.DBNAME) as connection:
            connection.row_factory = sqlite3.Row
            cursor = connection.cursor()
            cursor.execute(query, (from_user_id,))
            row = cursor.fetchone()
            return dict(row) if row else None


    def get_full_db_report(self, message=None):
        """
        Генерирует единый Excel-отчёт со всеми таблицами БД на разных листах
        
        Args:
            message: Опциональный объект сообщения от бота (для идентификации файла)
            
        Returns:
            tuple: (путь к файлу, словарь с количеством записей по таблицам)
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import pandas as pd
        import os

        # Создаем подключение к БД
        conn = create_connection(self.DBNAME)
        cursor = conn.cursor()
        
        # Получаем список всех таблиц (исключая системные)
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        tables = [table[0] for table in cursor.fetchall()]
        
        if not tables:
            return None, {}

        # Создаем директорию для отчетов
        os.makedirs("./reports", exist_ok=True)
        
        # Формируем имя файла
        file_id = message.from_user.id if message else "system"
        file_path = f"./reports/full_db_report_{file_id}.xlsx"
        
        # Создаем Excel-файл
        records_counts = {}
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for table in tables:
                try:
                    # Получаем данные таблицы
                    df = pd.read_sql_query(f"SELECT * FROM {table}", conn)
                    records_count = len(df)
                    records_counts[table] = records_count
                    
                    # Записываем данные в лист (ограничение длины имени листа до 31 символа)
                    sheet_name = table[:31]
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    
                    # Получаем объекты для форматирования
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Стили оформления
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    # Форматируем заголовки
                    for col_num, column_name in enumerate(df.columns, 1):
                        col_letter = get_column_letter(col_num)
                        
                        # Заголовок
                        header_cell = worksheet[f"{col_letter}1"]
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                        header_cell.alignment = Alignment(horizontal="center", vertical="center")
                        header_cell.border = thin_border
                        
                        # Автоподбор ширины столбца
                        max_length = max(
                            df[column_name].astype(str).str.len().max(),
                            len(str(column_name))
                        )
                        worksheet.column_dimensions[col_letter].width = min(max_length + 2, 30)
                    
                    # Форматируем ячейки с данными
                    for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
                        for cell in row:
                            cell.alignment = cell_alignment
                            cell.border = thin_border
                    
                    # Замораживаем заголовки и добавляем фильтры
                    worksheet.freeze_panes = "A2"
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                except Exception as e:
                    print(f"Ошибка при обработке таблицы {table}: {str(e)}")
                    continue
        
        conn.close()
        return file_path
